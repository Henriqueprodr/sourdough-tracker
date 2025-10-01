import os
import sys
import pathlib
import logging
import argparse
from rich import print
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import load_workbook, Workbook
from .commands import Config, Feeding

logging.basicConfig(
    filename="app.log",
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

filename = "starter_log.xlsx"

def parse_ratio(ratio_str: str) -> tuple[int, int, int]:
    try:
        parts = [int(x.strip()) for x in ratio_str.split(':')]
        if len(parts) != 3:
            logging.warning(f"Ratio must have 3 numbers in the format of starter:flour:water.")
            raise ValueError(f"Ratio must have 3 numbers in the format of starter:flour:water.")
        if any(p <= 0 for p in parts):
            logging.warning(f"All ratio values must be positive numbers.")
            raise ValueError(f"All ratio values must be positive numbers.")
        return tuple(parts)
    except ValueError as e:
        logging.error(f"Invalid ratio '{ratio_str}': {e}")
        raise ValueError(f"Invalid ratio '{ratio_str}': {e}")

def create_log_file(file: str = filename) -> None:
        if pathlib.Path(file).exists():
            logging.warning(f"Log file '{file}' already exists - skipping creation to avoid overwrite")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Feedings"
        headers = ["Date", "Jar Weight Total", "Starter Weight", "Target Weight",
        "Flour (calc)", "Water (calc)", "Smell", "Peak Hours", "Notes"]
        ws.append(headers)
        wb.save(file)
        logging.info(f"'{file}' file created.")

def append_feeding(feeding: Feeding, file: str = filename) -> None:
    logging.info(f"Attempting to append feeding for date {feeding.date}.")
    try:
        wb = load_workbook(file)
        logging.info(f"Loaded existing workbook: {file}")
    except FileNotFoundError:
        # If the file doesn't exist, create one
        logging.warning(f"Expected tracker file '{file}' not found - creating new tracker (possible first run or deleted file)")
        wb = Workbook()
        ws = wb.active
        ws.title = "Feedings"
        headers = ["Date", "Jar Weight Total", "Starter Weight", "Target Weight",
        "Flour (calc)", "Water (calc)", "Smell", "Peak Hours", "Notes"]
        ws.append(headers)
        wb.save(file)
    except PermissionError as e:
        # If the program lacks permission, exit the program and tell the user to close excel
        logging.error(f"Permission denied for '{file}': {e}")
        print(f"Cannot open '{file}'. Please close Excel and retry!")
        sys.exit(1)
    except InvalidFileException:
        # If the original file is corrupted, create a backup tracker
        logging.warning(f"File '{file} is corrupted. Renaming it to '{file}.bak' and creating a new tracker...")
        backup = file + ".bak"
        os.rename(file, backup)
        wb = Workbook()
        ws = wb.active
        ws.title = "Feedings"
        headers = ["Date", "Jar Weight Total", "Starter Weight", "Target Weight",
        "Flour (calc)", "Water (calc)", "Smell", "Peak Hours", "Notes"]
        ws.append(headers)
        wb.save(file)
    ws = wb.active
    ws.append(feeding.to_row())
    wb.save(file)
    logging.info(f"Successfully logged feeding on {feeding.date}")

def show_stats(file: str, limit: int=5) -> None:
    try:
        wb = load_workbook(file)
        logging.info(f"Successfully loaded workbook: {file}")
    except FileNotFoundError:
        logging.error(f"File does not exist. Run 'python main.py init' to set up your tracker.")
        sys.exit(1)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = rows[1:]

    # pick last N feedings
    for row in data[-limit:]:
        for header, value in zip(headers, row):
            print(f"{header}: {value}")
        print("---")

def main():
    parser = argparse.ArgumentParser(description="Sourdough Starter Tracker")
    subparsers = parser.add_subparsers(dest="command")

    init_parser = subparsers.add_parser("init", help="Set up the config")
    init_parser.add_argument("--jar-weight", type=int, required=True)
    init_parser.add_argument("--keep-target", type=int, required=True)
    init_parser.add_argument(
        "--ratio",
        type=str,
        default="1:2:2",
        help="Feeding ratio, format starter:flour:water (e.g., 1:2:2)"
    )
    
    feed_parser = subparsers.add_parser("feed", help="Log a feeding")
    feed_parser.add_argument("--weight", type=int, required=True)
    feed_parser.add_argument("--smell", type=str)
    feed_parser.add_argument("--peak-hours", type=int)
    feed_parser.add_argument("--notes", type=str)

    feed_parser = subparsers.add_parser("stats", help="Check on your starter")
    feed_parser.add_argument(
        "--limit",
        type=int,
        default=5,
        help="Number of recent feedings to show (default: 5)"
    )

    args = parser.parse_args()

    if args.command == "init":
        logging.info(f"User ran 'init' command with with jar_weight={args.jar_weight}, keep_target={args.keep_target}, and ratio={args.ratio}.")
        if args.jar_weight <= 0:
            logging.error("Jar weight must be a positive number.")
            sys.exit(1)

        if args.keep_target <= 0:
            logging.error("Keep target must be a positive number.")
            sys.exit(1)

        starter, flour, water = parse_ratio(args.ratio)

        config = Config(
            jar_weight=args.jar_weight,
            keep_target=args.keep_target,
            ratio=(starter, flour, water),
            path="."
        )
        config.save()

        create_log_file(filename)


    elif args.command == "feed":
        logging.info(f"User ran 'feed' command.")
        config = Config.load()
        feeding = Feeding(
            jar_weight_total=args.weight,
            config=config,
            smell=args.smell,
            peak_hours=args.peak_hours,
            notes=args.notes
        )
        print(f"Target jar weight after discard: [bold bright_cyan]{feeding.target_total_weight()}g")
        flour, water = feeding.flour_and_water()
        print(f"Add [bold bright_cyan]{flour}[/]g flour and [bold bright_cyan]{water}[/]g water")

        append_feeding(feeding, filename)

    elif args.command == "stats":
        logging.info(f"User ran 'stats' command.")
        show_stats(filename, limit=args.limit)

if __name__ == "__main__":
    main()