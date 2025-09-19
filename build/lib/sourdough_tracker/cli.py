from rich import print
import json
import argparse
import pathlib
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import date

filename = "starter_log.xlsx"

class Config():
    """Configure the script and save the user's information"""

    def __init__(self, jar_weight: int, keep_target, ratio, path):
        self.jar_weight = jar_weight
        self.keep_target = keep_target
        self.ratio = ratio
        self.path = path

    # Save the initial config.json file during configuration
    def save(self, filename="config.json"):
        with open(filename,"w+") as f:
            json.dump(self.__dict__, f, indent=4)

    # Load the config.json file
    @classmethod
    def load(cls, filename="config.json"):
        with open(filename, "r") as f:
            data = json.load(f)
            return cls(**data)
        
class Feeding():
    """Update and track the sourdough starter information"""

    def __init__(self, jar_weight_total, config: Config, smell, peak_hours, notes):
        self.date = date.today().isoformat()
        self.jar_weight_total = jar_weight_total
        self.config = config
        self.starter_weight = jar_weight_total - config.jar_weight
        self.smell = smell
        self.peak_hours = peak_hours
        self.notes = notes

    def target_weight(self):
        target_total_weight = self.config.jar_weight + self.config.keep_target
        return target_total_weight

    def flour_and_water(self) -> tuple[int, int]:
        flour = self.config.keep_target * self.config.ratio[1]
        water = self.config.keep_target * self.config.ratio[2]
        return flour, water
    
    def to_row(self):
        flour, water = self.flour_and_water()
        return [
            self.date,
            self.jar_weight_total,
            self.starter_weight,
            self.target_weight(),
            flour,
            water,
            self.smell,
            self.peak_hours,
            self.notes]
    
def create_log_file(file: str = filename):
        if pathlib.Path(file).exists():
            print("Log file already exists. Aborting to avoid overwrite.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Feedings"
        headers = ["Date", "Jar Weight Total", "Starter Weight", "Target Weight",
        "Flour (calc)", "Water (calc)", "Smell", "Peak Hours", "Notes"]
        ws.append(headers)
        wb.save(file)

def append_feeding(feeding, file: str = filename):
    wb = load_workbook(file)
    ws = wb.active
    ws.append(feeding.to_row())
    wb.save(file)
    print(f"Logged feeding on {feeding.date}")

def show_stats(file, limit=5):
    wb = load_workbook(file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = rows[1:]

    # pick last N feedings
    for row in data[-limit:]:
        for header, data in zip(headers, row):
            print(f"{header}: {data}")
        print("---")

def main():
    parser = argparse.ArgumentParser(description="Sourdough Starter Tracker")
    subparsers = parser.add_subparsers(dest="command")

    init_parser = subparsers.add_parser("init", help="Set up the config")
    init_parser.add_argument("--jar-weight", type=int, required=True)
    init_parser.add_argument("--keep-target", type=int, required=True)
    init_parser.add_argument(
        "--ratio",
        type=str,
        default="1:2:2",
        help="Feeding ratio, format starter:flour:water (e.g., 1:2:2)"
    )
    
    feed_parser = subparsers.add_parser("feed", help="Log a feeding")
    feed_parser.add_argument("--weight", type=int, required=True)
    feed_parser.add_argument("--smell", type=str)
    feed_parser.add_argument("--peak-hours", type=int)
    feed_parser.add_argument("--notes", type=str)

    feed_parser = subparsers.add_parser("stats", help="Check on your starter")
    feed_parser.add_argument(
        "--limit",
        type=int,
        default=5,
        help="Number of recent feedings to show (default: 5)"
    )

    args = parser.parse_args()

    if args.command == "init":
        # ugly ass way to split but i'll fix it later
        ratio = args.ratio.split(":")
        starter = int(ratio[0])
        flour = int(ratio[1])
        water = int(ratio[2])

        config = Config(
            jar_weight=args.jar_weight,
            keep_target=args.keep_target,
            ratio=(starter, flour, water),
            path="."
        )
        config.save()

        create_log_file(filename)

    elif args.command == "feed":
        config = Config.load()
        feeding = Feeding(
            jar_weight_total=args.weight,
            config=config,
            smell=args.smell,
            peak_hours=args.peak_hours,
            notes=args.notes
        )
        print(f"Target jar weight after discard: {feeding.target_weight()}g")
        flour, water = feeding.flour_and_water()
        print(f"Add [bright_cyan]{flour}[/]g flour and [bright_cyan]{water}[/]g water")

        append_feeding(feeding, filename)

    elif args.command == "stats":
        show_stats(filename, limit=args.limit)

if __name__ == "__main__":
    main()